from collections import Counter
from io import BytesIO

import frappe
from frappe.utils import add_days, nowdate
from openpyxl import load_workbook

from lifeline_tpa.lifeline_tpa.doctype.claims_po_batch.claims_po_batch import (
    import_claims,
    validate_po_file,
)
from lifeline_tpa.services.po_schema import MASTER_SHEET_COLUMNS


PILOT_FILE_NAME = "PO-pilot-5-providers.xlsx"
DEFAULT_PAYER = "DNIRC"
DEFAULT_PROVIDER_LIMIT = 5


def run(source_path, payer_name=DEFAULT_PAYER, provider_limit=DEFAULT_PROVIDER_LIMIT):
    """Create and import a small, repeatable PO pilot on the current site."""
    provider_limit = int(provider_limit)
    existing = _get_existing_pilot(payer_name, provider_limit)
    if existing:
        return existing

    workbook_content, providers = _build_pilot_workbook(source_path, provider_limit)
    company = _get_company()
    customer = _get_or_create_customer(payer_name)
    supplier_group = _get_or_create_supplier_group()

    for provider in providers:
        supplier = _get_or_create_supplier(provider["provider_name"], supplier_group)
        _get_or_create_mapping(
            provider["facility_id"],
            provider["provider_name"],
            supplier.name,
        )

    file_doc = frappe.get_doc(
        {
            "doctype": "File",
            "file_name": PILOT_FILE_NAME,
            "is_private": 1,
            "content": workbook_content,
        }
    ).insert(ignore_permissions=True)

    batch = frappe.get_doc(
        {
            "doctype": "Claims PO Batch",
            "company": company,
            "payer": customer.name,
            "posting_date": nowdate(),
            "expected_payment_date": add_days(nowdate(), 45),
            "source_file": file_doc.file_url,
        }
    ).insert(ignore_permissions=True)

    file_doc.db_set(
        {
            "attached_to_doctype": batch.doctype,
            "attached_to_name": batch.name,
            "attached_to_field": "source_file",
        }
    )

    preview = validate_po_file(batch.name)
    if not preview["valid"]:
        frappe.throw(
            "Pilot validation failed:\n"
            + "\n".join(error["message"] for error in preview["errors"])
        )

    result = import_claims(batch.name)
    frappe.db.commit()
    return {
        "status": "Imported",
        "batch_name": batch.name,
        "payer": customer.name,
        "company": company,
        "providers": providers,
        "total_providers": preview["total_providers"],
        "total_claims": result["imported_claims"],
        "total_amount": preview["total_amount"],
        "currency": preview["currency"],
        "external_po_ref": preview["external_po_ref"],
    }


def verify(batch_name):
    batch = frappe.get_doc("Claims PO Batch", batch_name)
    claim_totals = frappe.get_all(
        "Claim Record",
        filters={"po_batch": batch.name},
        fields=["count(name) as claim_count", "sum(claim_amount) as total_amount"],
    )[0]
    provider_totals = frappe.get_all(
        "Claim Record",
        filters={"po_batch": batch.name},
        fields=[
            "facility_id",
            "provider",
            "count(name) as claim_count",
            "sum(claim_amount) as total_amount",
        ],
        group_by="facility_id, provider",
        order_by="claim_count desc",
    )
    return {
        "batch_name": batch.name,
        "status": batch.status,
        "payer": batch.payer,
        "external_po_ref": batch.external_po_ref,
        "currency": batch.currency,
        "batch_claim_count": batch.total_claims,
        "database_claim_count": claim_totals.claim_count,
        "batch_total_amount": batch.total_amount,
        "database_total_amount": claim_totals.total_amount,
        "provider_count": len(provider_totals),
        "provider_totals": provider_totals,
    }


def _build_pilot_workbook(source_path, provider_limit):
    source = load_workbook(source_path, read_only=True, data_only=True)
    if "Master Sheet" not in source.sheetnames:
        frappe.throw("The source workbook does not contain the Master Sheet worksheet.")

    worksheet = source["Master Sheet"]
    rows = worksheet.iter_rows(values_only=True)
    headers = tuple(next(rows))
    if headers != MASTER_SHEET_COLUMNS:
        frappe.throw("The source workbook does not match the required PO columns.")

    facility_index = headers.index("Provider reference No")
    provider_index = headers.index("Provider Name")
    source_rows = list(rows)
    counts = Counter(
        str(row[facility_index]).strip()
        for row in source_rows
        if row[facility_index] is not None and str(row[facility_index]).strip()
    )
    selected_ids = [
        facility_id
        for facility_id, _count in sorted(
            counts.items(),
            key=lambda item: (-item[1], item[0]),
        )[:provider_limit]
    ]
    selected_set = set(selected_ids)

    output = load_workbook(source_path)
    output_sheet = output["Master Sheet"]
    output_sheet.delete_rows(2, output_sheet.max_row)
    provider_names = {}
    for row in source_rows:
        facility_id = str(row[facility_index]).strip()
        if facility_id not in selected_set:
            continue
        provider_names[facility_id] = str(row[provider_index]).strip()
        output_sheet.append(row)

    stream = BytesIO()
    output.save(stream)
    output.close()
    source.close()
    providers = [
        {
            "facility_id": facility_id,
            "provider_name": provider_names[facility_id],
            "claim_count": counts[facility_id],
        }
        for facility_id in selected_ids
    ]
    return stream.getvalue(), providers


def _get_company():
    company = frappe.db.get_value(
        "Company",
        {"default_currency": "AED", "name": ["not like", "_Test%"]},
        "name",
    )
    if not company:
        frappe.throw("No non-test AED company is configured.")
    return company


def _get_or_create_customer(customer_name):
    existing = frappe.db.get_value("Customer", {"customer_name": customer_name}, "name")
    if existing:
        return frappe.get_doc("Customer", existing)
    return frappe.get_doc(
        {
            "doctype": "Customer",
            "customer_name": customer_name,
            "customer_type": "Company",
            "customer_group": "Commercial",
            "territory": "United Arab Emirates",
        }
    ).insert(ignore_permissions=True)


def _get_or_create_supplier_group():
    group_name = "Healthcare Providers"
    if not frappe.db.exists("Supplier Group", group_name):
        frappe.get_doc(
            {
                "doctype": "Supplier Group",
                "supplier_group_name": group_name,
                "parent_supplier_group": "All Supplier Groups",
                "is_group": 0,
            }
        ).insert(ignore_permissions=True)
    return group_name


def _get_or_create_supplier(supplier_name, supplier_group):
    existing = frappe.db.get_value("Supplier", {"supplier_name": supplier_name}, "name")
    if existing:
        return frappe.get_doc("Supplier", existing)
    return frappe.get_doc(
        {
            "doctype": "Supplier",
            "supplier_name": supplier_name,
            "supplier_group": supplier_group,
            "supplier_type": "Company",
        }
    ).insert(ignore_permissions=True)


def _get_or_create_mapping(facility_id, provider_name, supplier):
    if frappe.db.exists("Provider Facility Mapping", facility_id):
        mapping = frappe.get_doc("Provider Facility Mapping", facility_id)
        if mapping.supplier != supplier:
            frappe.throw(
                f"Facility ID {facility_id} is already mapped to {mapping.supplier}."
            )
        return mapping
    return frappe.get_doc(
        {
            "doctype": "Provider Facility Mapping",
            "facility_id": facility_id,
            "provider_name_as_received": provider_name,
            "supplier": supplier,
        }
    ).insert(ignore_permissions=True)


def _get_existing_pilot(payer_name, provider_limit):
    batches = frappe.get_all(
        "Claims PO Batch",
        filters={
            "payer": payer_name,
            "external_po_ref": "DNIRC-000150/2026",
            "total_providers": provider_limit,
            "status": "Imported",
        },
        fields=[
            "name",
            "status",
            "external_po_ref",
            "currency",
            "total_providers",
            "total_claims",
            "total_amount",
        ],
        limit=1,
    )
    if not batches:
        return None
    batch = batches[0]
    batch["batch_name"] = batch.pop("name")
    batch["payer"] = payer_name
    batch["providers"] = frappe.get_all(
        "Provider Facility Mapping",
        filters={"disabled": 0},
        fields=["facility_id", "provider_name_as_received", "supplier"],
        limit=provider_limit,
        order_by="creation asc",
    )
    return batch
