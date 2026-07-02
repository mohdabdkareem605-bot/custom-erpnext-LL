from io import BytesIO

import frappe
from frappe.tests.utils import FrappeTestCase
from frappe.utils import nowdate
from openpyxl import Workbook, load_workbook

from lifeline_tpa.lifeline_tpa.doctype.claims_po_batch.claims_po_batch import (
    _build_processed_po_export,
    import_claims,
    process_bulk_claims,
    validate_po_file,
)
from lifeline_tpa.services.po_schema import MASTER_SHEET_COLUMNS, MASTER_SHEET_NAME
from lifeline_tpa.services.processed_po_export import PROCESSED_PO_COLUMNS


class TestClaimsPOBatch(FrappeTestCase):
    def test_validate_and_import_claim(self):
        suffix = frappe.generate_hash(length=8)
        company = frappe.get_all("Company", pluck="name", limit=1)[0]
        customer = _make_customer(f"Test Payer {suffix}")
        facility_id = f"FAC-{suffix}"
        claim_reference = f"CLAIM-{suffix}"
        source_file = _make_source_file(f"po-{suffix}.xlsx", facility_id, claim_reference)

        batch = frappe.get_doc(
            {
                "doctype": "Claims PO Batch",
                "company": company,
                "payer": customer.name,
                "posting_date": nowdate(),
                "source_file": source_file.file_url,
            }
        ).insert(ignore_permissions=True)

        first_preview = validate_po_file(batch.name)
        self.assertFalse(first_preview["valid"])
        self.assertIn(
            "missing_provider_mapping",
            [error["code"] for error in first_preview["errors"]],
        )

        supplier = _make_supplier(f"Test Provider {suffix}")
        frappe.get_doc(
            {
                "doctype": "Provider Facility Mapping",
                "facility_id": facility_id,
                "provider_name_as_received": supplier.supplier_name,
                "supplier": supplier.name,
            }
        ).insert(ignore_permissions=True)

        second_preview = validate_po_file(batch.name)
        self.assertTrue(second_preview["valid"])
        self.assertEqual(second_preview["total_claims"], 1)
        self.assertEqual(second_preview["total_amount"], 25.5)

        result = import_claims(batch.name)
        self.assertEqual(result["imported_claims"], 1)

        claim = frappe.get_doc("Claim Record", {"claim_reference": claim_reference})
        self.assertEqual(claim.po_batch, batch.name)
        self.assertEqual(claim.provider, supplier.name)
        self.assertEqual(claim.payer, customer.name)
        self.assertEqual(claim.claim_amount, 25.5)
        self.assertEqual(claim.payer_outstanding_amount, 25.5)
        self.assertEqual(claim.provider_outstanding_amount, 25.5)
        self.assertEqual(frappe.db.get_value("Claims PO Batch", batch.name, "status"), "Imported")

        batch.reload()
        batch.submit()
        batch.db_set("status", "Processing")
        report = process_bulk_claims(batch.name, requested_by="Administrator")
        self.assertEqual(report["status"], "Processed")

        claim.reload()
        self.assertTrue(claim.purchase_invoice)
        self.assertTrue(claim.purchase_invoice_item)
        self.assertTrue(claim.sales_invoice)
        self.assertTrue(claim.sales_invoice_item)
        self.assertEqual(
            frappe.db.get_value(
                "Purchase Invoice Item",
                claim.purchase_invoice_item,
                "rate",
            ),
            claim.claim_amount,
        )
        self.assertEqual(
            frappe.db.get_value(
                "Purchase Invoice",
                claim.purchase_invoice,
                ["disable_rounded_total", "grand_total", "outstanding_amount"],
            ),
            (1, claim.claim_amount, claim.claim_amount),
        )
        self.assertEqual(
            frappe.db.get_value(
                "Sales Invoice Item",
                claim.sales_invoice_item,
                "rate",
            ),
            claim.claim_amount,
        )
        self.assertEqual(
            frappe.db.get_value(
                "Sales Invoice",
                claim.sales_invoice,
                ["disable_rounded_total", "grand_total", "outstanding_amount"],
            ),
            (1, claim.claim_amount, claim.claim_amount),
        )

        export_content = _build_processed_po_export(batch.reload())
        export_workbook = load_workbook(BytesIO(export_content), read_only=True, data_only=True)
        export_sheet = export_workbook[MASTER_SHEET_NAME]
        export_rows = export_sheet.iter_rows(values_only=True)
        self.assertEqual(
            next(export_rows),
            (*MASTER_SHEET_COLUMNS, *PROCESSED_PO_COLUMNS),
        )
        exported_claim = next(export_rows)
        exported_values = dict(
            zip(
                (*MASTER_SHEET_COLUMNS, *PROCESSED_PO_COLUMNS),
                exported_claim,
                strict=True,
            )
        )
        self.assertEqual(exported_values["Unique Transaction ID"], claim_reference)
        self.assertEqual(exported_values["Payer"], customer.name)
        self.assertEqual(
            exported_values["Process Payable Number"],
            claim.purchase_invoice,
        )
        self.assertEqual(exported_values["Sales Invoice Number"], claim.sales_invoice)
        self.assertEqual(exported_values["Payer Paid Amount"], 0)
        self.assertEqual(exported_values["Payer Payment Status"], "Unpaid")
        self.assertEqual(exported_values["Provider Paid Amount"], 0)
        self.assertEqual(exported_values["Provider Payment Status"], "Unpaid")
        self.assertIsNone(next(export_rows, None))
        export_workbook.close()


def _make_customer(customer_name):
    return frappe.get_doc(
        {
            "doctype": "Customer",
            "customer_name": customer_name,
            "customer_type": "Company",
            "customer_group": "Commercial",
            "territory": "United Arab Emirates",
        }
    ).insert(ignore_permissions=True)


def _make_supplier(supplier_name):
    return frappe.get_doc(
        {
            "doctype": "Supplier",
            "supplier_name": supplier_name,
            "supplier_group": "Services",
            "supplier_type": "Company",
        }
    ).insert(ignore_permissions=True)


def _make_source_file(file_name, facility_id, claim_reference):
    values = {column: f"Value for {column}" for column in MASTER_SHEET_COLUMNS}
    values.update(
        {
            "Provider Name": "Test Provider",
            "Invoice #": "INV-1",
            "Payer Share CV": 25.5,
            "CV Currency": "AED",
            "POID/Year": "PO-TEST-1",
            "Provider reference No": facility_id,
            "Unique Transaction ID": claim_reference,
        }
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Master Sheet"
    worksheet.append(MASTER_SHEET_COLUMNS)
    worksheet.append([values[column] for column in MASTER_SHEET_COLUMNS])
    output = BytesIO()
    workbook.save(output)

    return frappe.get_doc(
        {
            "doctype": "File",
            "file_name": file_name,
            "is_private": 1,
            "content": output.getvalue(),
        }
    ).insert(ignore_permissions=True)
