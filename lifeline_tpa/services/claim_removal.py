import csv
import hashlib
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO

from openpyxl import load_workbook


REMOVAL_COLUMNS = (
    "Unique Transaction ID",
    "Removal Reason",
    "Approval Reference",
    "Removal Date",
)


@dataclass(frozen=True)
class ClaimRemovalRequest:
    row_number: int
    claim_reference: str
    removal_reason: str
    approval_reference: str
    removal_date: date


@dataclass(frozen=True)
class ParsedClaimRemoval:
    file_hash: str
    requests: tuple[ClaimRemovalRequest, ...]
    errors: tuple[dict, ...]

    @property
    def is_valid(self):
        return not self.errors and bool(self.requests)


def parse_claim_removal_file(content, file_name):
    rows = _read_rows(content, file_name)
    file_hash = hashlib.sha256(content).hexdigest()
    if not rows:
        return ParsedClaimRemoval(
            file_hash=file_hash,
            requests=(),
            errors=(_error(1, "empty_file", "The removal file is empty."),),
        )

    headers = tuple(_clean_text(value) for value in rows[0])
    if headers != REMOVAL_COLUMNS:
        return ParsedClaimRemoval(
            file_hash=file_hash,
            requests=(),
            errors=(
                _error(
                    1,
                    "invalid_headers",
                    "The file columns must exactly match: "
                    + ", ".join(REMOVAL_COLUMNS)
                    + ".",
                ),
            ),
        )

    requests = []
    errors = []
    seen = set()
    for row_number, row in enumerate(rows[1:], start=2):
        values = list(row) + [None] * (len(REMOVAL_COLUMNS) - len(row))
        if not any(_clean_text(value) for value in values):
            continue

        claim_reference = _clean_text(values[0])
        removal_reason = _clean_text(values[1])
        approval_reference = _clean_text(values[2])
        removal_date = _parse_date(values[3])

        if not claim_reference:
            errors.append(
                _error(
                    row_number,
                    "missing_claim_reference",
                    "Unique Transaction ID is required.",
                )
            )
        elif claim_reference in seen:
            errors.append(
                _error(
                    row_number,
                    "duplicate_claim_reference",
                    f"Claim {claim_reference} appears more than once in the file.",
                )
            )
        else:
            seen.add(claim_reference)

        if not removal_reason:
            errors.append(
                _error(row_number, "missing_reason", "Removal Reason is required.")
            )
        if not approval_reference:
            errors.append(
                _error(
                    row_number,
                    "missing_approval_reference",
                    "Approval Reference is required.",
                )
            )
        if not removal_date:
            errors.append(
                _error(
                    row_number,
                    "invalid_removal_date",
                    "Removal Date must be a valid date.",
                )
            )

        if (
            claim_reference
            and removal_reason
            and approval_reference
            and removal_date
        ):
            requests.append(
                ClaimRemovalRequest(
                    row_number=row_number,
                    claim_reference=claim_reference,
                    removal_reason=removal_reason,
                    approval_reference=approval_reference,
                    removal_date=removal_date,
                )
            )

    if not requests and not errors:
        errors.append(
            _error(2, "no_claims", "The removal file contains no claim rows.")
        )

    return ParsedClaimRemoval(
        file_hash=file_hash,
        requests=tuple(requests),
        errors=tuple(errors),
    )


def _read_rows(content, file_name):
    lower_name = file_name.lower()
    if lower_name.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = (
            workbook["Claim Removals"]
            if "Claim Removals" in workbook.sheetnames
            else workbook.active
        )
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        return rows
    if lower_name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return list(csv.reader(StringIO(text)))
    raise ValueError("The claim removal file must be an .xlsx or .csv file.")


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value)
    if not text:
        return None
    for date_format in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def _clean_text(value):
    return "" if value is None else str(value).strip()


def _error(row_number, code, message):
    return {
        "row_number": row_number,
        "code": code,
        "message": message,
    }
