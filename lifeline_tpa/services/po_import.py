import hashlib
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from lifeline_tpa.services.po_schema import (
    ACCOUNTING_AMOUNT_COLUMN,
    ACCOUNTING_CURRENCY_COLUMN,
    CLAIM_REFERENCE_COLUMN,
    EXTERNAL_PO_COLUMN,
    MASTER_SHEET_NAME,
    PROVIDER_FACILITY_COLUMN,
    PROVIDER_INVOICE_COLUMN,
    compare_headers,
)


REQUIRED_COLUMNS = (
    CLAIM_REFERENCE_COLUMN,
    EXTERNAL_PO_COLUMN,
    PROVIDER_FACILITY_COLUMN,
    ACCOUNTING_AMOUNT_COLUMN,
    ACCOUNTING_CURRENCY_COLUMN,
)


@dataclass(frozen=True)
class ValidationIssue:
    code: str
    message: str
    row_number: int | None = None
    field: str | None = None

    def as_dict(self):
        return {
            "code": self.code,
            "message": self.message,
            "row_number": self.row_number,
            "field": self.field,
        }


@dataclass(frozen=True)
class ParsedClaim:
    row_number: int
    claim_reference: str
    external_po_ref: str
    facility_id: str
    provider_name: str
    provider_invoice_number: str
    amount: Decimal
    currency: str
    source_data: dict


@dataclass
class ParsedWorkbook:
    file_hash: str
    claims: list[ParsedClaim] = field(default_factory=list)
    issues: list[ValidationIssue] = field(default_factory=list)

    @property
    def external_po_refs(self):
        return {claim.external_po_ref for claim in self.claims if claim.external_po_ref}

    @property
    def currencies(self):
        return {claim.currency for claim in self.claims if claim.currency}

    @property
    def facility_ids(self):
        return {claim.facility_id for claim in self.claims if claim.facility_id}

    @property
    def claim_references(self):
        return [claim.claim_reference for claim in self.claims if claim.claim_reference]

    @property
    def total_amount(self):
        return sum((claim.amount for claim in self.claims), Decimal("0"))


@dataclass
class ValidationResult:
    workbook: ParsedWorkbook
    provider_by_facility: dict[str, str]
    issues: list[ValidationIssue]

    @property
    def is_valid(self):
        return not self.issues

    def as_dict(self):
        po_refs = sorted(self.workbook.external_po_refs)
        currencies = sorted(self.workbook.currencies)
        return {
            "valid": self.is_valid,
            "file_hash": self.workbook.file_hash,
            "external_po_ref": po_refs[0] if len(po_refs) == 1 else None,
            "currency": currencies[0] if len(currencies) == 1 else None,
            "total_claims": len(self.workbook.claims),
            "total_providers": len(self.workbook.facility_ids),
            "total_amount": float(self.workbook.total_amount),
            "errors": [issue.as_dict() for issue in self.issues],
        }


def parse_workbook(content: bytes) -> ParsedWorkbook:
    file_hash = hashlib.sha256(content).hexdigest()
    parsed = ParsedWorkbook(file_hash=file_hash)

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError, KeyError) as exc:
        parsed.issues.append(
            ValidationIssue("invalid_workbook", f"The uploaded file is not a readable .xlsx workbook: {exc}")
        )
        return parsed

    try:
        if MASTER_SHEET_NAME not in workbook.sheetnames:
            parsed.issues.append(
                ValidationIssue(
                    "missing_master_sheet",
                    f'The workbook must contain a worksheet named "{MASTER_SHEET_NAME}".',
                )
            )
            return parsed

        worksheet = workbook[MASTER_SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, ())
        header_result = compare_headers(headers)
        if not header_result["valid"]:
            parsed.issues.extend(_header_issues(header_result))
            return parsed

        for row_number, values in enumerate(rows, start=2):
            if _is_blank_row(values):
                continue
            source_data = {
                header: _json_value(values[index]) for index, header in enumerate(header_result["expected"])
            }
            claim, row_issues = _parse_claim(row_number, source_data)
            parsed.issues.extend(row_issues)
            parsed.claims.append(claim)
    finally:
        workbook.close()

    parsed.issues.extend(_workbook_level_issues(parsed))
    return parsed


def validate_parsed_workbook(
    parsed: ParsedWorkbook,
    *,
    existing_claim_references=(),
    provider_by_facility=None,
    duplicate_file=False,
) -> ValidationResult:
    issues = list(parsed.issues)
    existing_claim_references = set(existing_claim_references)
    provider_by_facility = provider_by_facility or {}

    if duplicate_file:
        issues.append(
            ValidationIssue(
                "duplicate_file",
                "This exact workbook was already uploaded in another Claims PO Batch.",
            )
        )

    for claim_reference in sorted(existing_claim_references):
        issues.append(
            ValidationIssue(
                "existing_claim_reference",
                f"Claim reference {claim_reference} already exists in ERPNext.",
                field=CLAIM_REFERENCE_COLUMN,
            )
        )

    for facility_id in sorted(parsed.facility_ids - set(provider_by_facility)):
        issues.append(
            ValidationIssue(
                "missing_provider_mapping",
                f"Facility ID {facility_id} has no active Provider Facility Mapping.",
                field=PROVIDER_FACILITY_COLUMN,
            )
        )

    return ValidationResult(
        workbook=parsed,
        provider_by_facility=provider_by_facility,
        issues=issues,
    )


def _parse_claim(row_number, source_data):
    issues = []
    for column in REQUIRED_COLUMNS:
        if _is_blank(source_data[column]):
            issues.append(
                ValidationIssue(
                    "missing_required_value",
                    f"{column} is required.",
                    row_number=row_number,
                    field=column,
                )
            )

    amount = Decimal("0")
    raw_amount = source_data[ACCOUNTING_AMOUNT_COLUMN]
    if not _is_blank(raw_amount):
        try:
            amount = Decimal(str(raw_amount))
        except (InvalidOperation, ValueError):
            issues.append(
                ValidationIssue(
                    "invalid_amount",
                    f"{ACCOUNTING_AMOUNT_COLUMN} must be numeric.",
                    row_number=row_number,
                    field=ACCOUNTING_AMOUNT_COLUMN,
                )
            )
        else:
            if amount < 0:
                issues.append(
                    ValidationIssue(
                        "negative_amount",
                        f"{ACCOUNTING_AMOUNT_COLUMN} cannot be negative.",
                        row_number=row_number,
                        field=ACCOUNTING_AMOUNT_COLUMN,
                    )
                )

    return (
        ParsedClaim(
            row_number=row_number,
            claim_reference=_text(source_data[CLAIM_REFERENCE_COLUMN]),
            external_po_ref=_text(source_data[EXTERNAL_PO_COLUMN]),
            facility_id=_text(source_data[PROVIDER_FACILITY_COLUMN]),
            provider_name=_text(source_data["Provider Name"]),
            provider_invoice_number=_text(source_data[PROVIDER_INVOICE_COLUMN]),
            amount=amount,
            currency=_text(source_data[ACCOUNTING_CURRENCY_COLUMN]),
            source_data=source_data,
        ),
        issues,
    )


def _header_issues(header_result):
    issues = []
    if header_result["missing"]:
        issues.append(
            ValidationIssue(
                "missing_columns",
                "Missing columns: " + ", ".join(header_result["missing"]),
            )
        )
    if header_result["unexpected"]:
        issues.append(
            ValidationIssue(
                "unexpected_columns",
                "Unexpected columns: " + ", ".join(header_result["unexpected"]),
            )
        )
    if header_result["wrong_order"]:
        issues.append(
            ValidationIssue(
                "wrong_column_order",
                "The 34 Master Sheet columns are present but not in the required order.",
            )
        )
    return issues


def _workbook_level_issues(parsed):
    issues = []
    if not parsed.claims and not parsed.issues:
        issues.append(ValidationIssue("empty_workbook", "Master Sheet contains no claim rows."))

    if len(parsed.external_po_refs) > 1:
        issues.append(
            ValidationIssue(
                "multiple_po_references",
                "The workbook contains multiple POID/Year values: "
                + ", ".join(sorted(parsed.external_po_refs)),
                field=EXTERNAL_PO_COLUMN,
            )
        )

    if len(parsed.currencies) > 1:
        issues.append(
            ValidationIssue(
                "multiple_currencies",
                "The workbook contains multiple CV Currency values: "
                + ", ".join(sorted(parsed.currencies)),
                field=ACCOUNTING_CURRENCY_COLUMN,
            )
        )

    claim_counts = {}
    for claim in parsed.claims:
        if not claim.claim_reference:
            continue
        claim_counts.setdefault(claim.claim_reference, []).append(claim.row_number)
    for claim_reference, row_numbers in sorted(claim_counts.items()):
        if len(row_numbers) > 1:
            issues.append(
                ValidationIssue(
                    "duplicate_claim_reference",
                    f"Claim reference {claim_reference} is repeated on rows "
                    + ", ".join(str(row_number) for row_number in row_numbers)
                    + ".",
                    field=CLAIM_REFERENCE_COLUMN,
                )
            )
    return issues


def _json_value(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _text(value):
    return "" if value is None else str(value).strip()


def _is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _is_blank_row(values):
    return not any(not _is_blank(value) for value in values)
