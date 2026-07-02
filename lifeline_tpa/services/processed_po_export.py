import json
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lifeline_tpa.services.po_schema import MASTER_SHEET_COLUMNS, MASTER_SHEET_NAME


PROCESSED_PO_COLUMNS = (
    "Payer",
    "Process Payable Number",
    "Sales Invoice Number",
    "Payer Paid Amount",
    "Payer Payment Reference",
    "Payer Payment Date",
    "Payer Payment Status",
    "Provider Paid Amount",
    "Provider Payment Reference",
    "Provider Payment Date",
    "Provider Payment Status",
)

PROCESSED_PO_FIELD_BY_COLUMN = {
    "Payer": "payer",
    "Process Payable Number": "process_payable_number",
    "Sales Invoice Number": "sales_invoice",
    "Payer Paid Amount": "payer_allocated_amount",
    "Payer Payment Reference": "payer_payment_reference",
    "Payer Payment Date": "payer_payment_date",
    "Payer Payment Status": "payer_settlement_status",
    "Provider Paid Amount": "provider_paid_amount",
    "Provider Payment Reference": "provider_payment_reference",
    "Provider Payment Date": "provider_payment_date",
    "Provider Payment Status": "provider_settlement_status",
}


def build_processed_po_workbook(claims):
    claims = list(claims)
    if not claims:
        raise ValueError("The batch has no Claim Records to export.")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = MASTER_SHEET_NAME
    headers = (*MASTER_SHEET_COLUMNS, *PROCESSED_PO_COLUMNS)
    worksheet.append(headers)

    for claim in sorted(claims, key=lambda row: int(_value(row, "source_row_number") or 0)):
        source_data = _source_data(claim)
        row = [_excel_value(source_data.get(column)) for column in MASTER_SHEET_COLUMNS]
        row.extend(
            _excel_value(_value(claim, PROCESSED_PO_FIELD_BY_COLUMN[column]))
            for column in PROCESSED_PO_COLUMNS
        )
        worksheet.append(row)

    _format_worksheet(worksheet, len(headers), len(claims) + 1)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _source_data(claim):
    raw_source_data = _value(claim, "source_data_json")
    try:
        source_data = json.loads(raw_source_data or "{}")
    except (TypeError, json.JSONDecodeError) as exc:
        raise ValueError(
            f"Claim {_value(claim, 'claim_reference')} contains invalid source data."
        ) from exc

    missing_columns = [column for column in MASTER_SHEET_COLUMNS if column not in source_data]
    if missing_columns:
        raise ValueError(
            f"Claim {_value(claim, 'claim_reference')} is missing original source columns: "
            + ", ".join(missing_columns)
        )
    return source_data


def _value(row, fieldname):
    if isinstance(row, dict):
        return row.get(fieldname)
    return getattr(row, fieldname, None)


def _excel_value(value):
    if isinstance(value, (date, datetime)):
        return value
    if value is None:
        return ""
    return value


def _format_worksheet(worksheet, column_count, row_count):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{row_count}"
    worksheet.row_dimensions[1].height = 36

    for index, header in enumerate((*MASTER_SHEET_COLUMNS, *PROCESSED_PO_COLUMNS), start=1):
        width = max(14, min(32, len(header) + 3))
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for column in ("Payer Paid Amount", "Provider Paid Amount"):
        column_index = len(MASTER_SHEET_COLUMNS) + PROCESSED_PO_COLUMNS.index(column) + 1
        for row_index in range(2, row_count + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = "#,##0.00"
