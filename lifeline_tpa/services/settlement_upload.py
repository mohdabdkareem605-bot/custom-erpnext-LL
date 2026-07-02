import csv
import hashlib
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

from openpyxl import load_workbook


PAYER_RECEIPT_COLUMNS = (
    "claim_unique_number",
    "amount_paid",
    "lifeline_bank_account",
)

PROVIDER_PAYMENT_COLUMNS = (
    "claim_unique_number",
    "amount_paid",
    "lifeline_bank_account",
    "payment_reference",
)


@dataclass(frozen=True)
class SettlementUploadRow:
    row_number: int
    claim_reference: str
    amount_paid: Decimal
    lifeline_bank_account: str
    payment_reference: str | None = None


@dataclass(frozen=True)
class ParsedSettlementUpload:
    file_hash: str
    rows: tuple[SettlementUploadRow, ...]
    errors: tuple[dict, ...]

    @property
    def is_valid(self):
        return not self.errors and bool(self.rows)


def parse_payer_receipt_file(content, file_name):
    return _parse_settlement_file(
        content=content,
        file_name=file_name,
        expected_columns=PAYER_RECEIPT_COLUMNS,
        requires_payment_reference=False,
        label="payer receipt",
    )


def parse_provider_payment_file(content, file_name):
    return _parse_settlement_file(
        content=content,
        file_name=file_name,
        expected_columns=PROVIDER_PAYMENT_COLUMNS,
        requires_payment_reference=True,
        label="provider payment",
    )


def _parse_settlement_file(
    *,
    content,
    file_name,
    expected_columns,
    requires_payment_reference,
    label,
):
    rows = _read_rows(content, file_name)
    file_hash = hashlib.sha256(content).hexdigest()
    if not rows:
        return ParsedSettlementUpload(
            file_hash=file_hash,
            rows=(),
            errors=(_error(1, "empty_file", f"The {label} file is empty."),),
        )

    headers = tuple(_clean_text(value) for value in rows[0])
    if headers != expected_columns:
        return ParsedSettlementUpload(
            file_hash=file_hash,
            rows=(),
            errors=(
                _error(
                    1,
                    "invalid_headers",
                    "The file columns must exactly match: "
                    + ", ".join(expected_columns)
                    + ".",
                ),
            ),
        )

    parsed_rows = []
    errors = []
    seen = set()
    for row_number, row in enumerate(rows[1:], start=2):
        values = list(row) + [None] * (len(expected_columns) - len(row))
        if not any(_clean_text(value) for value in values):
            continue

        claim_reference = _clean_text(values[0])
        amount_paid = _parse_amount(values[1])
        lifeline_bank_account = _clean_text(values[2])
        payment_reference = _clean_text(values[3]) if requires_payment_reference else None

        if not claim_reference:
            errors.append(
                _error(
                    row_number,
                    "missing_claim_reference",
                    "claim_unique_number is required.",
                )
            )
        elif claim_reference in seen:
            errors.append(
                _error(
                    row_number,
                    "duplicate_claim_reference",
                    f"Claim {claim_reference} appears more than once in the file.",
                )
            )
        else:
            seen.add(claim_reference)

        if amount_paid is None:
            errors.append(
                _error(
                    row_number,
                    "invalid_amount_paid",
                    "amount_paid must be a valid number.",
                )
            )
        elif amount_paid <= Decimal("0"):
            errors.append(
                _error(
                    row_number,
                    "non_positive_amount_paid",
                    "amount_paid must be greater than zero.",
                )
            )

        if not lifeline_bank_account:
            errors.append(
                _error(
                    row_number,
                    "missing_lifeline_bank_account",
                    "lifeline_bank_account is required.",
                )
            )

        if requires_payment_reference and not payment_reference:
            errors.append(
                _error(
                    row_number,
                    "missing_payment_reference",
                    "payment_reference is required.",
                )
            )

        if (
            claim_reference
            and amount_paid is not None
            and amount_paid > Decimal("0")
            and lifeline_bank_account
            and (payment_reference or not requires_payment_reference)
        ):
            parsed_rows.append(
                SettlementUploadRow(
                    row_number=row_number,
                    claim_reference=claim_reference,
                    amount_paid=amount_paid,
                    lifeline_bank_account=lifeline_bank_account,
                    payment_reference=payment_reference,
                )
            )

    if not parsed_rows and not errors:
        errors.append(
            _error(2, "no_claims", f"The {label} file contains no claim rows.")
        )

    return ParsedSettlementUpload(
        file_hash=file_hash,
        rows=tuple(parsed_rows),
        errors=tuple(errors),
    )


def _read_rows(content, file_name):
    lower_name = file_name.lower()
    if lower_name.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        return rows
    if lower_name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return list(csv.reader(StringIO(text)))
    raise ValueError("The settlement file must be an .xlsx or .csv file.")


def _parse_amount(value):
    text = _clean_text(value)
    if not text:
        return None
    try:
        return Decimal(text.replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def _clean_text(value):
    return "" if value is None else str(value).strip()


def _error(row_number, code, message):
    return {
        "row_number": row_number,
        "code": code,
        "message": message,
    }
