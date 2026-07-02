import calendar
import hashlib
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


DEFAULT_CURRENCY = "AED"

ENDORSEMENT_SHEET_REQUIRED_COLUMNS = (
    "EndorseYear",
    "EndorseMonth",
    "payer",
    "GroupName",
    "PolicyNo",
    "PolicyStartDate",
    "PolicyStopDate",
    "CardNo",
    "MemberID",
    "EffDate_EndosDate",
    "EndorsementType",
    "TPA/Service Fees",
)

OPTIONAL_COLUMNS = (
    "EmployeeID",
    "MemberName",
    "MemberSecondname",
    "Policy Category",
    "Gross Premium",
    "Prorata Gross Premium",
)

VALID_ENDORSEMENT_TYPES = ("Inception", "Addition", "Deletion")


@dataclass(frozen=True)
class RevenueValidationIssue:
    code: str
    message: str
    row_number: int | None = None
    field: str | None = None

    def as_dict(self):
        return {
            "code": self.code,
            "message": self.message,
            "row_number": self.row_number,
            "field": self.field,
        }


@dataclass(frozen=True)
class ParsedRevenueEvent:
    row_number: int
    endorsement_year: int
    endorsement_month: int
    payer_name: str
    group_name: str
    policy_no: str
    policy_start_date: date
    policy_stop_date: date
    employee_id: str
    member_name: str
    member_second_name: str
    card_no: str
    member_id: str
    endorsement_date: date
    endorsement_type: str
    tpa_fee: Decimal
    policy_category: str
    gross_premium: Decimal
    prorata_gross_premium: Decimal
    currency: str
    source_data: dict

    @property
    def is_deletion(self):
        return self.endorsement_type == "Deletion"

    @property
    def document_type(self):
        return "Credit Note" if self.is_deletion else "Sales Invoice"

    @property
    def event_key(self):
        return (
            self.member_id,
            self.policy_no,
            self.endorsement_date.isoformat() if self.endorsement_date else "",
            self.endorsement_type,
            str(self.tpa_fee),
        )


@dataclass
class ParsedRevenueWorkbook:
    file_hash: str
    events: list[ParsedRevenueEvent] = field(default_factory=list)
    issues: list[RevenueValidationIssue] = field(default_factory=list)

    @property
    def payer_names(self):
        return {event.payer_name for event in self.events if event.payer_name}

    @property
    def member_ids(self):
        return [event.member_id for event in self.events if event.member_id]

    @property
    def endorsement_periods(self):
        return {
            (event.endorsement_year, event.endorsement_month)
            for event in self.events
            if event.endorsement_year and event.endorsement_month
        }

    @property
    def currencies(self):
        return {event.currency for event in self.events if event.currency}

    @property
    def total_positive_fee(self):
        return sum((event.tpa_fee for event in self.events if event.tpa_fee > 0), Decimal("0"))

    @property
    def total_credit_fee(self):
        return sum((event.tpa_fee for event in self.events if event.tpa_fee < 0), Decimal("0"))

    @property
    def net_fee(self):
        return sum((event.tpa_fee for event in self.events), Decimal("0"))


@dataclass
class RevenueValidationResult:
    workbook: ParsedRevenueWorkbook
    customer_by_payer_name: dict[str, str]
    issues: list[RevenueValidationIssue]

    @property
    def is_valid(self):
        return not self.issues

    def as_dict(self):
        periods = sorted(self.workbook.endorsement_periods)
        currencies = sorted(self.workbook.currencies)
        return {
            "valid": self.is_valid,
            "file_hash": self.workbook.file_hash,
            "endorsement_year": periods[0][0] if len(periods) == 1 else None,
            "endorsement_month": periods[0][1] if len(periods) == 1 else None,
            "currency": currencies[0] if len(currencies) == 1 else DEFAULT_CURRENCY,
            "total_events": len(self.workbook.events),
            "total_payers": len(self.workbook.payer_names),
            "total_positive_events": sum(not event.is_deletion for event in self.workbook.events),
            "total_deletion_events": sum(event.is_deletion for event in self.workbook.events),
            "total_invoice_amount": float(self.workbook.total_positive_fee),
            "total_credit_note_amount": float(self.workbook.total_credit_fee),
            "net_tpa_fee": float(self.workbook.net_fee),
            "errors": [issue.as_dict() for issue in self.issues],
        }


def parse_revenue_workbook(content: bytes) -> ParsedRevenueWorkbook:
    parsed = ParsedRevenueWorkbook(file_hash=hashlib.sha256(content).hexdigest())

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError, KeyError) as exc:
        parsed.issues.append(
            RevenueValidationIssue(
                "invalid_workbook",
                f"The uploaded endorsement report is not a readable .xlsx workbook: {exc}",
            )
        )
        return parsed

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        headers = [_clean_header(value) for value in next(rows, ())]
        header_map = {header: index for index, header in enumerate(headers) if header}

        missing_columns = [
            column for column in ENDORSEMENT_SHEET_REQUIRED_COLUMNS if column not in header_map
        ]
        if missing_columns:
            parsed.issues.append(
                RevenueValidationIssue(
                    "missing_required_columns",
                    "The endorsement report is missing required columns: "
                    + ", ".join(missing_columns),
                    row_number=1,
                )
            )
            return parsed

        for row_number, values in enumerate(rows, start=2):
            if _is_blank_row(values):
                continue
            source_data = {
                column: _json_value(_get_value(values, header_map, column))
                for column in (*ENDORSEMENT_SHEET_REQUIRED_COLUMNS, *OPTIONAL_COLUMNS)
                if column in header_map
            }
            event, row_issues = _parse_event(row_number, source_data)
            parsed.issues.extend(row_issues)
            parsed.events.append(event)
    finally:
        workbook.close()

    parsed.issues.extend(_workbook_level_issues(parsed))
    return parsed


def validate_revenue_workbook(
    parsed: ParsedRevenueWorkbook,
    *,
    customer_by_payer_name=None,
    duplicate_file=False,
    existing_event_keys=(),
) -> RevenueValidationResult:
    issues = list(parsed.issues)
    customer_by_payer_name = customer_by_payer_name or {}
    existing_event_keys = set(existing_event_keys)

    if duplicate_file:
        issues.append(
            RevenueValidationIssue(
                "duplicate_file",
                "This exact endorsement report was already uploaded in another TPA Revenue Batch.",
            )
        )

    if len(parsed.endorsement_periods) > 1:
        issues.append(
            RevenueValidationIssue(
                "multiple_endorsement_months",
                "One TPA Revenue Batch can contain only one endorsement month.",
            )
        )

    for payer_name in sorted(parsed.payer_names - set(customer_by_payer_name)):
        issues.append(
            RevenueValidationIssue(
                "unknown_payer",
                f"Payer {payer_name} is not mapped to an ERPNext Customer.",
                field="payer",
            )
        )

    for event in parsed.events:
        if event.event_key in existing_event_keys:
            issues.append(
                RevenueValidationIssue(
                    "existing_revenue_event",
                    (
                        f"Member {event.member_id} / policy {event.policy_no} / "
                        f"{event.endorsement_type} on {event.endorsement_date} already exists."
                    ),
                    row_number=event.row_number,
                    field="MemberID",
                )
            )

    return RevenueValidationResult(
        workbook=parsed,
        customer_by_payer_name=customer_by_payer_name,
        issues=issues,
    )


def _parse_event(row_number, source_data):
    issues = []
    for column in ENDORSEMENT_SHEET_REQUIRED_COLUMNS:
        if _is_blank(source_data.get(column)):
            issues.append(
                RevenueValidationIssue(
                    "missing_required_value",
                    f"{column} is required.",
                    row_number=row_number,
                    field=column,
                )
            )

    endorsement_year = _parse_year(source_data.get("EndorseYear"), row_number, issues)
    endorsement_month = _parse_month(source_data.get("EndorseMonth"), row_number, issues)
    policy_start_date = _parse_date(
        source_data.get("PolicyStartDate"),
        "PolicyStartDate",
        row_number,
        issues,
    )
    policy_stop_date = _parse_date(
        source_data.get("PolicyStopDate"),
        "PolicyStopDate",
        row_number,
        issues,
    )
    endorsement_date = _parse_date(
        source_data.get("EffDate_EndosDate"),
        "EffDate_EndosDate",
        row_number,
        issues,
    )
    endorsement_type = _parse_endorsement_type(
        source_data.get("EndorsementType"),
        row_number,
        issues,
    )
    tpa_fee = _parse_decimal(source_data.get("TPA/Service Fees"), "TPA/Service Fees", row_number, issues)
    gross_premium = _parse_optional_decimal(source_data.get("Gross Premium"))
    prorata_gross_premium = _parse_optional_decimal(source_data.get("Prorata Gross Premium"))

    if policy_start_date and policy_stop_date and policy_stop_date <= policy_start_date:
        issues.append(
            RevenueValidationIssue(
                "invalid_policy_period",
                "PolicyStopDate must be after PolicyStartDate.",
                row_number=row_number,
                field="PolicyStopDate",
            )
        )
    if (
        policy_start_date
        and policy_stop_date
        and endorsement_date
        and not (policy_start_date <= endorsement_date <= policy_stop_date)
    ):
        issues.append(
            RevenueValidationIssue(
                "endorsement_date_outside_policy_period",
                "EffDate_EndosDate must fall inside the policy period.",
                row_number=row_number,
                field="EffDate_EndosDate",
            )
        )
    if endorsement_type in ("Inception", "Addition") and tpa_fee <= 0:
        issues.append(
            RevenueValidationIssue(
                "invalid_tpa_fee_sign",
                f"{endorsement_type} rows must have a positive TPA/Service Fees amount.",
                row_number=row_number,
                field="TPA/Service Fees",
            )
        )
    if endorsement_type == "Deletion" and tpa_fee >= 0:
        issues.append(
            RevenueValidationIssue(
                "invalid_tpa_fee_sign",
                "Deletion rows must have a negative TPA/Service Fees amount.",
                row_number=row_number,
                field="TPA/Service Fees",
            )
        )

    return (
        ParsedRevenueEvent(
            row_number=row_number,
            endorsement_year=endorsement_year,
            endorsement_month=endorsement_month,
            payer_name=_string(source_data.get("payer")),
            group_name=_string(source_data.get("GroupName")),
            policy_no=_string(source_data.get("PolicyNo")),
            policy_start_date=policy_start_date,
            policy_stop_date=policy_stop_date,
            employee_id=_string(source_data.get("EmployeeID")),
            member_name=_string(source_data.get("MemberName")),
            member_second_name=_string(source_data.get("MemberSecondname")),
            card_no=_string(source_data.get("CardNo")),
            member_id=_string(source_data.get("MemberID")),
            endorsement_date=endorsement_date,
            endorsement_type=endorsement_type,
            tpa_fee=tpa_fee,
            policy_category=_string(source_data.get("Policy Category")),
            gross_premium=gross_premium,
            prorata_gross_premium=prorata_gross_premium,
            currency=DEFAULT_CURRENCY,
            source_data=source_data,
        ),
        issues,
    )


def _workbook_level_issues(parsed):
    issues = []
    grouped = {}
    for event in parsed.events:
        if not event.member_id or not event.endorsement_year or not event.endorsement_month:
            continue
        key = (event.member_id, event.endorsement_year, event.endorsement_month)
        grouped.setdefault(key, []).append(event)

    for (member_id, year, month), events in grouped.items():
        if len(events) == 1:
            continue
        valid_same_month_delete = (
            len(events) == 2
            and sum(event.endorsement_type == "Deletion" for event in events) == 1
            and sum(event.endorsement_type in ("Inception", "Addition") for event in events) == 1
        )
        if valid_same_month_delete:
            continue

        issues.append(
            RevenueValidationIssue(
                "invalid_duplicate_member",
                (
                    f"Member {member_id} appears {len(events)} times in "
                    f"{calendar.month_name[month]} {year}. Repeats are allowed only for "
                    "one Inception/Addition and one Deletion in the same month."
                ),
                field="MemberID",
            )
        )
    return issues


def _parse_year(value, row_number, issues):
    try:
        year = int(str(value).strip())
        if year < 1900 or year > 2200:
            raise ValueError
        return year
    except (TypeError, ValueError):
        issues.append(
            RevenueValidationIssue(
                "invalid_year",
                "EndorseYear must be a valid year.",
                row_number=row_number,
                field="EndorseYear",
            )
        )
        return 0


def _parse_month(value, row_number, issues):
    text = _string(value)
    try:
        month = int(text)
        if 1 <= month <= 12:
            return month
    except ValueError:
        pass

    normalized = text.strip().lower()
    month_lookup = {
        name.lower(): index for index, name in enumerate(calendar.month_name) if name
    }
    month_lookup.update(
        {name.lower(): index for index, name in enumerate(calendar.month_abbr) if name}
    )
    if normalized in month_lookup:
        return month_lookup[normalized]

    issues.append(
        RevenueValidationIssue(
            "invalid_month",
            "EndorseMonth must be a valid month number or month name.",
            row_number=row_number,
            field="EndorseMonth",
        )
    )
    return 0


def _parse_date(value, field, row_number, issues):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _string(value)
    for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    issues.append(
        RevenueValidationIssue(
            "invalid_date",
            f"{field} must be a valid date.",
            row_number=row_number,
            field=field,
        )
    )
    return None


def _parse_endorsement_type(value, row_number, issues):
    text = _string(value)
    for valid_type in VALID_ENDORSEMENT_TYPES:
        if text.lower() == valid_type.lower():
            return valid_type
    issues.append(
        RevenueValidationIssue(
            "invalid_endorsement_type",
            "EndorsementType must be Inception, Addition, or Deletion.",
            row_number=row_number,
            field="EndorsementType",
        )
    )
    return text


def _parse_decimal(value, field, row_number, issues):
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, AttributeError, ValueError):
        issues.append(
            RevenueValidationIssue(
                "invalid_amount",
                f"{field} must be a valid number.",
                row_number=row_number,
                field=field,
            )
        )
        return Decimal("0")


def _parse_optional_decimal(value):
    if _is_blank(value):
        return Decimal("0")
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, AttributeError, ValueError):
        return Decimal("0")


def _get_value(values, header_map, column):
    index = header_map.get(column)
    if index is None or index >= len(values):
        return None
    return values[index]


def _clean_header(value):
    return _string(value).strip()


def _string(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_blank(value):
    return value is None or str(value).strip() == ""


def _is_blank_row(values):
    return not values or all(_is_blank(value) for value in values)


def _json_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value
