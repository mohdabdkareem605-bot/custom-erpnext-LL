import json
import unittest
from io import BytesIO

from openpyxl import load_workbook

from lifeline_tpa.services.po_schema import MASTER_SHEET_COLUMNS, MASTER_SHEET_NAME
from lifeline_tpa.services.processed_po_export import (
    PROCESSED_PO_COLUMNS,
    build_processed_po_workbook,
)


class TestProcessedPOExport(unittest.TestCase):
    def test_builds_original_and_processing_columns_in_source_order(self):
        claims = [
            _claim("CLAIM-2", 3, "PINV-2", "SINV-2"),
            _claim("CLAIM-1", 2, "PINV-1", "SINV-1"),
        ]

        content = build_processed_po_workbook(claims)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook[MASTER_SHEET_NAME]
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()

        self.assertEqual(rows[0], (*MASTER_SHEET_COLUMNS, *PROCESSED_PO_COLUMNS))
        self.assertEqual(rows[1][MASTER_SHEET_COLUMNS.index("Unique Transaction ID")], "CLAIM-1")
        self.assertEqual(rows[2][MASTER_SHEET_COLUMNS.index("Unique Transaction ID")], "CLAIM-2")

        process_payable_index = len(MASTER_SHEET_COLUMNS) + PROCESSED_PO_COLUMNS.index(
            "Process Payable Number"
        )
        sales_invoice_index = len(MASTER_SHEET_COLUMNS) + PROCESSED_PO_COLUMNS.index(
            "Sales Invoice Number"
        )
        self.assertEqual(rows[1][process_payable_index], "PINV-1")
        self.assertEqual(rows[1][sales_invoice_index], "SINV-1")

    def test_rejects_claim_with_incomplete_original_source_data(self):
        claim = _claim("CLAIM-1", 2, "PINV-1", "SINV-1")
        claim["source_data_json"] = json.dumps({"Unique Transaction ID": "CLAIM-1"})

        with self.assertRaisesRegex(ValueError, "missing original source columns"):
            build_processed_po_workbook([claim])


def _claim(reference, row_number, purchase_invoice, sales_invoice):
    source_data = {column: f"Value for {column}" for column in MASTER_SHEET_COLUMNS}
    source_data["Unique Transaction ID"] = reference
    source_data["Payer Share CV"] = 25.5
    return {
        "claim_reference": reference,
        "source_row_number": row_number,
        "source_data_json": json.dumps(source_data),
        "payer": "Test Payer",
        "process_payable_number": purchase_invoice,
        "sales_invoice": sales_invoice,
        "payer_allocated_amount": 10,
        "payer_payment_reference": "BANK-1",
        "payer_payment_date": "2026-06-22",
        "payer_settlement_status": "Partially Paid",
        "provider_paid_amount": 5,
        "provider_payment_reference": "CHEQUE-1",
        "provider_payment_date": "2026-06-23",
        "provider_settlement_status": "Partially Paid",
    }


if __name__ == "__main__":
    unittest.main()
