import unittest
from pathlib import Path

from openpyxl import load_workbook

from lifeline_tpa.services.claim_removal import REMOVAL_COLUMNS
from lifeline_tpa.services.po_schema import MASTER_SHEET_COLUMNS, MASTER_SHEET_NAME
from lifeline_tpa.services.settlement_upload import (
    PAYER_RECEIPT_COLUMNS,
    PROVIDER_PAYMENT_COLUMNS,
)


REPO_ROOT = Path(__file__).resolve().parents[2]
PUBLIC_FILES = REPO_ROOT / "lifeline_tpa" / "public" / "files"
DOCTYPE_DIR = REPO_ROOT / "lifeline_tpa" / "lifeline_tpa" / "doctype"


class TestUploadTemplates(unittest.TestCase):
    def test_claims_po_template_matches_expected_headers(self):
        headers = _read_headers("claims_po_upload_template.xlsx", MASTER_SHEET_NAME)
        self.assertEqual(headers, MASTER_SHEET_COLUMNS)

    def test_payer_receipt_template_matches_expected_headers(self):
        headers = _read_headers("payer_receipt_upload_template.xlsx")
        self.assertEqual(headers, PAYER_RECEIPT_COLUMNS)

    def test_provider_payment_template_matches_expected_headers(self):
        headers = _read_headers("provider_payment_upload_template.xlsx")
        self.assertEqual(headers, PROVIDER_PAYMENT_COLUMNS)

    def test_claim_removal_template_matches_expected_headers(self):
        headers = _read_headers("claim_removal_upload_template.xlsx", "Claim Removals")
        self.assertEqual(headers, REMOVAL_COLUMNS)

    def test_upload_forms_link_to_public_templates(self):
        cases = (
            (
                "claims_po_batch/claims_po_batch.js",
                "claims_po_upload_template.xlsx",
            ),
            (
                "payer_receipt_settlement/payer_receipt_settlement.js",
                "payer_receipt_upload_template.xlsx",
            ),
            (
                "provider_payment_settlement/provider_payment_settlement.js",
                "provider_payment_upload_template.xlsx",
            ),
            (
                "claim_removal_batch/claim_removal_batch.js",
                "claim_removal_upload_template.xlsx",
            ),
        )

        for script, template_name in cases:
            with self.subTest(script=script):
                script_text = (DOCTYPE_DIR / script).read_text()
                self.assertIn(
                    f"/assets/lifeline_tpa/files/{template_name}", script_text
                )


def _read_headers(template_name, sheet_name=None):
    workbook = load_workbook(PUBLIC_FILES / template_name, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        return tuple(cell.value for cell in next(sheet.iter_rows(max_row=1)))
    finally:
        workbook.close()
